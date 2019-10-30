from openpyxl import load_workbook, Workbook

def sortByLength(pumsaList):
  maxLength=0
  resultList = []
  for pumsa in pumsaList:
    if pumsa is None: continue
    if maxLength < len(pumsa):
      maxLength=len(pumsa)
  for pLength in range(maxLength):
    tempList = []
    for pumsa in pumsaList:
      if pumsa is None: continue
      if len(pumsa) == pLength:
        tempList.append(pumsa)
    tempList.sort()
    resultList.extend(tempList)
  return resultList

def findVerb():
  ALL = che_yeon['A']
  ADJ = che_yeon['B']
  temp_set = set()
  for item in ALL: temp_set.add(item.value)
  for item in ADJ: temp_set - {item}
  temp_set = temp_set- {'같은 것들'}
  temp_set = temp_set - {'형용사'}
  temp_list = list(temp_set)
  temp_list = sortByLength(temp_list)
  idx = 2
  for item in temp_list:
    che_yeon.cell(row=idx, column=3).value = item
    idx += 1



load_verb = load_workbook("verb.xlsx", data_only=True)
load_adj = load_workbook("adjective.xlsx", data_only=True)
load_che_yeon = load_workbook("che_yeon.xlsx", data_only=True)

verb = load_verb['Sheet1']
adj = load_adj['Sheet1']
che_yeon = load_che_yeon['Sheet1']
col_v = verb['D']
col_a = adj['D']
verb_set = set()
adj_set = set()

for cell in col_v: verb_set.add(cell.value)
for cell in col_a: adj_set.add(cell.value)

same_set = verb_set & adj_set
same_list = sortByLength(same_set)

idx = 2
for item in same_list:
    che_yeon.cell(row=idx, column=1).value = item
    idx += 1
che_yeon['A1'] = '같은 것들'

#findVerb()
load_che_yeon.save('che_yeon.xlsx')
