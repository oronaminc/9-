from openpyxl import load_workbook, Workbook

def preprocess(load_ws, text):
    global colA
    temp_set = set()
    temp_set2 = set()
    temp_int, temp_int2 = 2, 2
    for cell in colA:
        eomi = cell.value
        if(eomi.count('(')): eomi = eomi[:-4]
        
        idx = eomi.find('-')
        if idx != -1:
          #temp = eomi[:idx]
          #if len(eomi[:idx]) == 1:
          temp = eomi[:idx]+(jaem[(ord(eomi[idx+1])- 0xAC00) // 21 // 28])
          #print(jaem[(ord(eomi[idx+1])- 0xAC00) // 21 // 28])
          temp_set2.add(temp)
        else: temp_set2.add(eomi[:-2]+find_none_consonant(eomi[-2]))

        eomi=eomi.replace("-","")
        load_ws.cell(row=cell.row, column=2).value = eomi
        temp_set.add(eomi[-2:])
        
    load_ws['B1'] = text

    temp_list = list(temp_set)
    temp_list.sort()
    for item in temp_list:
        load_ws.cell(row=temp_int, column=3).value = item
        temp_int += 1
    load_ws['C1'] = text+' 체언판별'

    temp_list2 = list(temp_set2)
    temp_list2 = sortByLength(temp_list2)
    print(temp_list2)
    for item in temp_list2:
        load_ws.cell(row=temp_int2, column=4).value = item
        temp_int2 += 1
    load_ws['D1'] = text+' 어간판별'

def sortByLength(pumsaList):
  maxLength=0
  resultList = []
  for pumsa in pumsaList:
    if pumsa is None: continue
    if maxLength < len(pumsa):
      maxLength=len(pumsa)
  for pLength in range(maxLength):
    tempList = []
    for pumsa in pumsaList:
      if pumsa is None: continue
      if len(pumsa) == pLength:
        tempList.append(pumsa)
    tempList.sort()
    resultList.extend(tempList)
  return resultList

def find_none_consonant(word):
    temp_ord = ord(word)
    if temp_ord%28 > 16:
        temp_ord -= ((temp_ord%28)-16)
    elif temp_ord%28 < 16:
        temp_ord -= ((temp_ord%28)+12)
    return(chr(temp_ord))

def del_same(load_ws, word):
  global colAA
  colAA = L_same['C']
  colD = load_ws['D']
  temp_set = set()
  for item in colD: temp_set.add(item.value)
  temp_set = temp_set - {'형용사 어간판별'}
  for item in colAA: temp_set -= {item.value}
  temp_set = temp_set - {word}
  idx = 2
  temp_list = list(temp_set)
  temp_list = sortByLength(temp_list)
  for item in temp_list:
    load_ws.cell(row=idx, column=5).value = item
    idx += 1
  load_ws['E1'] = word + ' 중복제거'

jaem = "ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ"
load_wb = load_workbook("adjective.xlsx", data_only=True)
load_same = load_workbook("che_yeon.xlsx", data_only=True)
load_ws = load_wb['Sheet1']
L_same = load_same['Sheet1']
colA = load_ws['A']
colAA = L_same['C']
preprocess(load_ws, '형용사')
del_same(load_ws, '동사')
load_wb.save('adjective.xlsx')
