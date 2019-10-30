# 형용사와 동사를 추가한 어미를 구분하는 함수

from openpyxl import load_workbook, Workbook


def preprocess(load_ws, text):
    global colA
    global except_case
    global except_case_list
    for cell in colA:
        eomi = cell.value
        #print(eomi)
        eomi=eomi.replace("-","")
        if(eomi.count('(')): eomi = eomi[:-4]
        load_ws.cell(row=cell.row, column=5).value = eomi
        for item in except_case_list:
            if eomi.count(item):
                if eomi[1:] != '':
                    except_case[item].add(eomi[1:])
    load_ws['E1'] = text

def find_prob(load_ws, text):
    global colA2
    global colE
    global real_set
    first = True
    for cell2 in colA2:
        if first:
            first = False
            continue
        temp = cell2.value
        min_idx = 99999
        if find_except(temp): continue
        else:
            for cell in colE:
                n = temp.find(cell.value)
                if n != -1:
                    if min_idx > n: min_idx = n
            temp = temp[:min_idx]
            real_set.add(temp)
    
    idx = 2
    sheet2 = load_wb2.create_sheet('sheet2')
    sheet2.title = text
    sheet2['A1'] = text
    real_list = list(real_set)
    real_list.sort()
    for item in real_list:
        if item == '': continue
        sheet2.cell(row=idx, column=1).value = item
        idx += 1
    
    load_wb2.save('verb.xlsx')


    #print(real_set)

def find_except(word):
    global except_case
    global except_case_list
    global real_set
    for key in except_case_list:
        for value in except_case[key]:
            n = len(value)
            #print('뭐니?', word[-n:])
            if value == word[-n:]:
                try: temp = ord(word[-(n+1)])%28
                except: break
                if temp == 20:
                    real_set.add(word[:-(n+1)]+chr(ord(word[-(n+1)])-4))
                    return True
                elif temp == 24:
                    real_set.add(word[:-(n+1)]+chr(ord(word[-(n+1)])-8))
                    return True
                elif temp == 4:
                    real_set.add(word[:-(n+1)]+chr(ord(word[-(n+1)])-16))
                    return True
                elif temp == 5:
                    real_set.add(word[:-(n+1)]+chr(ord(word[-(n+1)])-17))
                    return True
    return False


load_wb = load_workbook("eomi.xlsx", data_only=True)
load_wb2 = load_workbook("verb.xlsx", data_only=True)
load_ws = load_wb['Raw']
load_ws2 = load_wb2['Sheet1']
colA = load_ws['A']
colE = load_ws['E']
colA2 = load_ws2['B']
except_case = {'ㄴ':set(), 'ㄹ':set(), 'ㅁ':set(), 'ㅂ':set()}
except_case_list = ['ㄴ', 'ㄹ', 'ㅁ', 'ㅂ']
real_set = set()
preprocess(load_ws, '순수 어미')
find_prob(load_ws, '순수 어간_동사')

print(real_set)

load_wb.save('eomi.xlsx')
