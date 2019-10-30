#전체 어미를 찾는 함수

from openpyxl import load_workbook, Workbook

load_wb = load_workbook("eomi.xlsx", data_only=True)
load_ws = load_wb['순수_어미']
colA = load_ws['A']
start = True
eomi_set = set()
eomi_idx = 2
for item in colA:
    eomi_set.add(item.value)

eomi_set = eomi_set-{'순수 어미'}
eomi_list = list(eomi_set)
eomi_list.sort()
print(eomi_list)
for item in eomi_list:
    load_ws.cell(row=eomi_idx, column=2).value = item
    eomi_idx += 1

load_ws['B1'] = '순수 어미2'
load_wb.save('eomi.xlsx')

